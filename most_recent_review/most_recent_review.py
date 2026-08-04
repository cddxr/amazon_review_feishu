#!/usr/bin/env python3
"""通过 ASIN 字典抓取 Amazon 商品按 Most recent 排序的评论。

示例：
    py most_recent_review.py
    py most_recent_review.py B0XXXXXXXX --limit 20
    py most_recent_review.py B0XXXXXXXX --domain amazon.co.uk --headless

Amazon 可能要求登录或显示验证码。默认使用可复用的 Chrome 用户目录并显示
浏览器；遇到验证时，请在浏览器中手动完成，然后回到终端按 Enter。
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from deep_translator import GoogleTranslator
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# 在这里填写要抓取的 ASIN。左边是 ASIN，右边是便于识别的商品备注名。
# 直接运行 `py most_recent_review.py` 时，会依次抓取字典中的全部商品。
ASIN_MAP = {
    "B0FRSFSJGD": "",
    "B0FRSJVZZH": "",
    "B0FRSHR4CP": "",
    "B0FRSL6Y82": "",
}

ASIN_PATTERN = re.compile(r"^[A-Z0-9]{10}$")
DOMAIN_PATTERN = re.compile(r"^(?:www\.)?amazon\.[a-z.]{2,10}$", re.IGNORECASE)
REVIEW_SELECTOR = "[data-hook='review']"
SHOW_MORE_SELECTOR = "[data-hook='show-more-button']"
ENABLE_TRANSLATE = True
TRANSLATION_CACHE: dict[str, str] = {}
TRANSLATOR = GoogleTranslator(source="auto", target="zh-CN")


@dataclass
class Review:
    asin: str
    product_name: str
    review_id: str
    title: str
    title_zh: str
    body: str
    body_zh: str
    rating: str
    author: str
    date: str
    verified_purchase: bool
    helpful_votes: str
    url: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="读取 ASIN_MAP 字典，抓取 Amazon 按 Most recent 排序的最新评论。"
    )
    parser.add_argument("asin", nargs="?", help="10 位 Amazon ASIN")
    parser.add_argument("--limit", type=int, default=110, help="最多保存几条评论（默认 110）")
    parser.add_argument(
        "--max-clicks",
        type=int,
        default=10,
        help="最多点击几次 Show 10 more reviews（默认 10）",
    )
    parser.add_argument(
        "--domain", default="amazon.com", help="Amazon 站点域名（默认 amazon.com）"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(__file__).resolve().parent / "output",
        help="输出目录",
    )
    parser.add_argument(
        "--profile-dir",
        type=Path,
        default=Path(__file__).resolve().parent / ".chrome-profile",
        help="Chrome 登录状态保存目录",
    )
    parser.add_argument("--headless", action="store_true", help="无界面运行 Chrome")
    parser.add_argument(
        "--no-pause",
        action="store_true",
        help="遇到登录/验证码页面时直接报错，不等待手动处理",
    )
    return parser.parse_args()


def normalize_asin(raw_asin: str) -> str:
    asin = raw_asin.strip().upper()
    if not ASIN_PATTERN.fullmatch(asin):
        raise ValueError("ASIN 必须是 10 位字母或数字，例如 B0XXXXXXXX。")
    return asin


def normalize_domain(raw_domain: str) -> str:
    domain = raw_domain.strip().lower()
    domain = re.sub(r"^https?://", "", domain).strip("/")
    domain = re.sub(r"^www\.", "", domain)
    if not DOMAIN_PATTERN.fullmatch(domain):
        raise ValueError("domain 必须是 Amazon 域名，例如 amazon.com 或 amazon.co.uk。")
    return domain


def build_review_url(domain: str, asin: str) -> str:
    """生成 Amazon Portal 的 Most recent 评论区链接。"""
    return (
        f"https://www.{domain}/portal/customer-reviews/{asin}/"
        "ref=cm_cr_arp_d_viewopt_fmt?"
        "ie=UTF8"
        "&filterByStar=all_stars"
        "&reviewerType=all_reviews"
        "&sortBy=recent"
        "&formatType=current_format"
        "#reviews-filter-bar"
    )


def find_chromedriver() -> Path | None:
    """优先使用手动指定或 Selenium 缓存中的 ChromeDriver。"""
    configured_path = os.getenv("CHROMEDRIVER_PATH", "").strip()
    if configured_path:
        driver_path = Path(configured_path).expanduser()
        if not driver_path.is_file():
            raise FileNotFoundError(
                f"CHROMEDRIVER_PATH 指向的文件不存在：{driver_path}"
            )
        return driver_path

    cache_root = Path.home() / ".cache" / "selenium" / "chromedriver" / "win64"
    candidates = list(cache_root.glob("*/chromedriver.exe"))
    if not candidates:
        return None

    def version_key(path: Path) -> tuple[int, ...]:
        try:
            return tuple(int(part) for part in path.parent.name.split("."))
        except ValueError:
            return (0,)

    return max(candidates, key=version_key)


def build_driver(profile_dir: Path, headless: bool) -> webdriver.Chrome:
    profile_dir.mkdir(parents=True, exist_ok=True)
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument(f"--user-data-dir={profile_dir.resolve()}")
    options.add_argument("--window-size=1440,1000")
    options.add_argument("--lang=en-US")
    options.add_argument("--disable-notifications")

    driver_path = find_chromedriver()
    if driver_path:
        print(f"使用本地 ChromeDriver：{driver_path}")
        driver = webdriver.Chrome(
            service=Service(executable_path=str(driver_path)),
            options=options,
        )
    else:
        print("本地没有找到 ChromeDriver，改用 Selenium Manager。")
        driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(45)
    return driver


def page_needs_human(driver: webdriver.Chrome) -> bool:
    source = driver.page_source.lower()
    current_url = driver.current_url.lower()
    markers = (
        "robot check",
        "enter the characters you see below",
        "type the characters you see in this image",
        "signin",
        "authportal",
    )
    return (
        "/errors/validatecaptcha" in current_url
        or "/ap/signin" in current_url
        or any(marker in source for marker in markers[:3])
    )


def wait_for_review_page(
    driver: webdriver.Chrome, allow_manual_pause: bool, timeout: int = 20
) -> bool:
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, REVIEW_SELECTOR)
            or page_needs_human(d)
        )
    except TimeoutException:
        return False

    if page_needs_human(driver):
        if not allow_manual_pause:
            raise RuntimeError("Amazon 要求登录或验证码验证。")
        if not sys.stdin.isatty():
            raise RuntimeError("Amazon 要求人工验证，但当前终端不能交互。")
        print("Amazon 要求登录或验证。请在 Chrome 中完成后，回到终端按 Enter...")
        input()
        try:
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, REVIEW_SELECTOR))
            )
        except TimeoutException as exc:
            raise RuntimeError("验证后仍未检测到评论，请重试。") from exc
    return bool(driver.find_elements(By.CSS_SELECTOR, REVIEW_SELECTOR))


def first_text(element, *selectors: str) -> str:
    for selector in selectors:
        matches = element.find_elements(By.CSS_SELECTOR, selector)
        if matches:
            return " ".join(matches[0].text.split())
    return ""


def extract_rating(text: str) -> str:
    if not text:
        return ""
    match = re.search(
        r"([1-5](?:[.,]\d)?)\s*(?:(?:out of|von|sur)\s*5|(?:颗|顆)?星)",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        match = re.search(r"\b([1-5](?:[.,]\d)?)\b", text)
    if not match:
        return ""
    return str(int(float(match.group(1).replace(",", "."))))


def extract_review_rating(card) -> str:
    """兼容 Portal 页面中可见、隐藏及无障碍属性形式的评分。"""
    selectors = (
        "[data-hook='review-star-rating'] .a-icon-alt",
        "[data-hook='cmps-review-star-rating'] .a-icon-alt",
        "[data-hook='review-star-rating']",
        "[data-hook='cmps-review-star-rating']",
        ".review-rating .a-icon-alt",
        ".review-rating",
    )
    attributes = ("aria-label", "title", "innerText", "textContent")

    for selector in selectors:
        for element in card.find_elements(By.CSS_SELECTOR, selector):
            candidates = [element.text]
            candidates.extend(element.get_attribute(name) for name in attributes)
            for candidate in candidates:
                rating = extract_rating(candidate or "")
                if rating:
                    return rating

    # 某些 Portal 版本只把评分写进评论标题链接的 aria-label。
    for element in card.find_elements(By.CSS_SELECTOR, "[data-hook='review-title']"):
        for name in attributes:
            rating = extract_rating(element.get_attribute(name) or "")
            if rating:
                return rating

    # 最后一层兜底只匹配明确的“x out of 5”结构，避免误把日期当评分。
    card_text = card.get_attribute("textContent") or ""
    match = re.search(
        r"([1-5](?:[.,]\d)?)\s*(?:out of|von|sur)\s*5",
        card_text,
        flags=re.IGNORECASE,
    )
    return str(int(float(match.group(1).replace(",", ".")))) if match else ""


def clean_title(title: str) -> str:
    # 某些站点会把“5.0 out of 5 stars”放进标题元素。
    return re.sub(
        r"^\s*[0-5](?:[.,]\d)?\s+(?:out of|von|sur)\s+5\s+stars?\s*",
        "",
        title,
        flags=re.IGNORECASE,
    ).strip()


def translate_to_chinese(text: str) -> str:
    """翻译单条或长文本；接口返回空值时重试，失败后返回空字符串。"""
    if not ENABLE_TRANSLATE or not text:
        return ""
    if text in TRANSLATION_CACHE:
        return TRANSLATION_CACHE[text]

    # Google 翻译单次文本存在长度限制，长评论分段处理。
    chunks = [text[index : index + 4500] for index in range(0, len(text), 4500)]
    translated_chunks: list[str] = []
    for chunk in chunks:
        for attempt in range(1, 3):
            try:
                translated = TRANSLATOR.translate(chunk)
                if not isinstance(translated, str) or not translated.strip():
                    raise RuntimeError("Google 翻译接口返回了空结果")
                translated_chunks.append(translated)
                break
            except Exception as exc:
                if attempt == 2:
                    print(f"Google 翻译失败，保留英文原文并继续：{exc}")
                    return ""
                time.sleep(1)

    result = "\n".join(translated_chunks)
    TRANSLATION_CACHE[text] = result
    return result


def translate_unique_texts(texts: list[str], batch_size: int = 20) -> dict[str, str]:
    """完整文本去重后批量翻译，批量失败的项目再单独重试。"""
    unique_texts = list(dict.fromkeys(text for text in texts if text))
    pending = [text for text in unique_texts if text not in TRANSLATION_CACHE]

    # 超长文本由单条翻译方法负责分段；短文本使用批量接口。
    short_texts = [text for text in pending if len(text) <= 4500]
    long_texts = [text for text in pending if len(text) > 4500]

    print(
        f"翻译去重：原始 {len(texts)} 段，唯一 {len(unique_texts)} 段，"
        f"缓存命中 {len(unique_texts) - len(pending)} 段"
    )

    for start in range(0, len(short_texts), batch_size):
        batch = short_texts[start : start + batch_size]
        print(
            f"正在批量翻译 {start + 1}-"
            f"{min(start + len(batch), len(short_texts))}/{len(short_texts)}..."
        )
        try:
            results = TRANSLATOR.translate_batch(batch)
        except Exception as exc:
            print(f"批量翻译失败，将逐条重试：{exc}")
            results = []
        if not isinstance(results, (list, tuple)):
            print("批量翻译接口返回空结果，将逐条重试。")
            results = []

        for index, text in enumerate(batch):
            translated = results[index] if index < len(results) else None
            if isinstance(translated, str) and translated.strip():
                TRANSLATION_CACHE[text] = translated
            else:
                # 只重试批量结果中失败或返回 None 的文本。
                translate_to_chinese(text)

    for text in long_texts:
        translate_to_chinese(text)

    return {text: TRANSLATION_CACHE.get(text, "") for text in unique_texts}


def translate_reviews(reviews: list[Review]) -> None:
    """一次性翻译评论列表的标题和正文，直接回填 Review 对象。"""
    if not ENABLE_TRANSLATE or not reviews:
        return

    texts = [text for review in reviews for text in (review.title, review.body) if text]
    translations = translate_unique_texts(texts)
    for review in reviews:
        review.title_zh = translations.get(review.title, "")
        review.body_zh = translations.get(review.body, "")


def find_show_more_button(driver: webdriver.Chrome, timeout: int = 15):
    """滚动到页面底部并寻找可见的 Show more 按钮。"""
    end_time = time.time() + timeout
    while time.time() < end_time:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        for button in driver.find_elements(By.CSS_SELECTOR, SHOW_MORE_SELECTOR):
            try:
                if button.is_displayed() and button.is_enabled():
                    return button
            except WebDriverException:
                continue

        # data-hook 变化时，按英文按钮文本兜底。
        xpath = (
            "//button[contains(normalize-space(.), 'Show 10 more reviews') "
            "or contains(normalize-space(.), 'Show more reviews')]"
        )
        for button in driver.find_elements(By.XPATH, xpath):
            try:
                if button.is_displayed() and button.is_enabled():
                    return button
            except WebDriverException:
                continue
        time.sleep(1)
    return None


def click_show_more(driver: webdriver.Chrome, max_clicks: int) -> int:
    """点击 Show 10 more reviews，并等待评论数量确实增加。"""
    successful_clicks = 0
    for click_number in range(1, max_clicks + 1):
        before_count = len(driver.find_elements(By.CSS_SELECTOR, REVIEW_SELECTOR))
        print(f"准备第 {click_number} 次点击，当前有 {before_count} 条评论...")

        button = find_show_more_button(driver)
        if button is None:
            print("未找到 Show 10 more reviews 按钮，停止继续加载。")
            break

        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});", button
            )
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", button)
            WebDriverWait(driver, 30).until(
                lambda d: len(d.find_elements(By.CSS_SELECTOR, REVIEW_SELECTOR))
                > before_count
            )
        except (TimeoutException, WebDriverException) as exc:
            print(f"第 {click_number} 次点击后未加载出新评论：{exc}")
            break

        after_count = len(driver.find_elements(By.CSS_SELECTOR, REVIEW_SELECTOR))
        successful_clicks += 1
        print(f"第 {click_number} 次点击成功：{before_count} -> {after_count} 条")
        time.sleep(1)

    return successful_clicks


def extract_reviews(
    driver: webdriver.Chrome, domain: str, asin: str, product_name: str
) -> list[Review]:
    reviews: list[Review] = []
    cards = driver.find_elements(By.CSS_SELECTOR, REVIEW_SELECTOR)
    for card in cards:
        review_id = (card.get_attribute("id") or "").strip()
        title = clean_title(first_text(card, "[data-hook='review-title']"))
        body = first_text(card, "[data-hook='review-body']")
        reviews.append(
            Review(
                asin=asin,
                product_name=product_name,
                review_id=review_id,
                title=title,
                title_zh="",
                body=body,
                body_zh="",
                rating=extract_review_rating(card),
                author=first_text(card, ".a-profile-name"),
                date=first_text(card, "[data-hook='review-date']"),
                verified_purchase=bool(
                    card.find_elements(By.CSS_SELECTOR, "[data-hook='avp-badge']")
                ),
                helpful_votes=first_text(
                    card, "[data-hook='helpful-vote-statement']"
                ),
                url=(
                    f"https://www.{domain}/gp/customer-reviews/{review_id}"
                    if review_id
                    else ""
                ),
            )
        )
    return reviews


def scrape_reviews(
    driver: webdriver.Chrome,
    domain: str,
    asin: str,
    product_name: str,
    limit: int,
    max_clicks: int,
    allow_manual_pause: bool,
) -> list[Review]:
    url = build_review_url(domain, asin)
    print(f"正在读取 Portal 评论区：{url}")
    driver.get(url)

    if not wait_for_review_page(driver, allow_manual_pause):
        raise RuntimeError("Portal 评论区没有加载出评论，可能需要登录或页面结构已变化。")

    clicks = click_show_more(driver, max_clicks)
    print(f"Show more 实际成功点击 {clicks} 次")

    collected: list[Review] = []
    seen: set[str] = set()
    for review in extract_reviews(driver, domain, asin, product_name):
        key = review.review_id or f"{review.author}|{review.date}|{review.body}"
        if key in seen:
            continue
        seen.add(key)
        collected.append(review)
        if len(collected) >= limit:
            break

    print(f"Portal 评论区共提取 {len(collected)} 条评论")
    translate_reviews(collected)
    return collected


def save_xlsx(reviews: list[Review], xlsx_path: Path) -> Path:
    """把评论列表保存为带基础格式的 XLSX 文件。"""
    fieldnames = list(Review.__dataclass_fields__)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Most Recent Reviews"
    worksheet.append(fieldnames)

    for review in reviews:
        row = asdict(review)
        worksheet.append([row.get(field, "") for field in fieldnames])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = {
        "A": 16,  # asin
        "B": 24,  # product_name
        "C": 24,  # review_id
        "D": 45,  # title
        "E": 45,  # title_zh
        "F": 80,  # body
        "G": 80,  # body_zh
        "H": 10,  # rating
        "I": 20,  # author
        "J": 35,  # date
        "K": 18,  # verified_purchase
        "L": 24,  # helpful_votes
        "M": 70,  # url
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 让评论链接在 Excel 中可以直接点击。
    url_column = fieldnames.index("url") + 1
    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=url_column)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"

    workbook.save(xlsx_path)
    return xlsx_path


def save_results(
    reviews: list[Review], asin: str, output_dir: Path
) -> tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"{asin}_most_recent_{timestamp}.csv"
    json_path = output_dir / f"{asin}_most_recent_{timestamp}.json"
    xlsx_path = output_dir / f"{asin}_most_recent_{timestamp}.xlsx"
    rows = [asdict(review) for review in reviews]
    fieldnames = list(Review.__dataclass_fields__)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    with json_path.open("w", encoding="utf-8") as file:
        json.dump(rows, file, ensure_ascii=False, indent=2)

    save_xlsx(reviews, xlsx_path)
    return csv_path, json_path, xlsx_path


def main() -> int:
    args = parse_args()

    try:
        domain = normalize_domain(args.domain)
        if args.limit <= 0 or args.max_clicks < 0:
            raise ValueError("--limit 必须大于 0，--max-clicks 不能小于 0。")

        if args.asin:
            targets = [(normalize_asin(args.asin), args.asin.strip().upper())]
        else:
            if not ASIN_MAP:
                raise ValueError("ASIN_MAP 不能为空，请先在脚本顶部填写 ASIN。")
            targets = [
                (normalize_asin(asin), str(product_name).strip() or asin)
                for asin, product_name in ASIN_MAP.items()
            ]
    except ValueError as exc:
        print(f"参数错误：{exc}", file=sys.stderr)
        return 2

    driver: webdriver.Chrome | None = None
    failed = 0
    try:
        driver = build_driver(args.profile_dir, args.headless)
        for asin, product_name in targets:
            print(f"\n开始抓取：{product_name}（{asin}）")
            try:
                reviews = scrape_reviews(
                    driver=driver,
                    domain=domain,
                    asin=asin,
                    product_name=product_name,
                    limit=args.limit,
                    max_clicks=args.max_clicks,
                    allow_manual_pause=not args.no_pause and not args.headless,
                )
                if not reviews:
                    raise RuntimeError("没有抓取到评论。")
                csv_path, json_path, xlsx_path = save_results(
                    reviews, asin, args.output_dir
                )
                print(f"完成：抓取 {len(reviews)} 条 Most recent 评论")
                print(f"CSV ：{csv_path.resolve()}")
                print(f"JSON：{json_path.resolve()}")
                print(f"XLSX：{xlsx_path.resolve()}")
            except (RuntimeError, TimeoutException, WebDriverException) as exc:
                failed += 1
                print(f"{asin} 抓取失败：{exc}", file=sys.stderr)
        return 1 if failed else 0
    except (RuntimeError, TimeoutException, WebDriverException) as exc:
        print(f"抓取失败：{exc}", file=sys.stderr)
        return 1
    finally:
        if driver is not None:
            driver.quit()


if __name__ == "__main__":
    raise SystemExit(main())
